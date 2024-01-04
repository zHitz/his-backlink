import os
import requests
import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

current_datetime = datetime.now().strftime("%Y-%m-%d")

# Tạo đường dẫn đến thư mục /his-backlink/
base_directory = os.path.abspath("./")
# Mở file Excel
file_path = os.path.join(base_directory, f'merged_data.xlsx')

# Kiểm tra xem file đã tồn tại hay chưa
if not os.path.exists(file_path):
    # Tạo DataFrame rỗng
    empty_df = pd.DataFrame()

    # Lưu DataFrame vào file Excel
    empty_df.to_excel(file_path, index=False)
    
# Đổi tên file cũ thành 'merged_data.xlsx'
old_filename = 'merged_data.xlsx'
os.rename(old_filename, 'temp_data.xlsx')


# Đọc dữ liệu từ file results_backlink.xlsx và temp_data.xlsx
data1 = pd.read_excel(f'results_backlink_{current_datetime}.xlsx')
data2 = pd.read_excel('temp_data.xlsx')

# Thêm dữ liệu từ data2 vào data1
merged_data = pd.concat([data1, data2], ignore_index=True)

# Đổi tên file mới thành 'merged_data.xlsx'
new_filename = 'merged_data.xlsx'
merged_data.to_excel(new_filename, index=False)

os.remove('temp_data.xlsx')


# Lấy ngày giờ hiện tại
now = datetime.now()

# Kiểm tra nếu là thứ 5 và đã qua 6 giờ tối
if now.weekday() == 3 and now.hour >= 18:

    # Tạo đường dẫn đến thư mục /his-backlink/
    base_directory = os.path.abspath("./")

    # Mở file Excel
    file_path = os.path.join(base_directory, f'merged_data.xlsx')
    print(file_path)

    #### Lọc Duplicate ####
    df = pd.read_excel(file_path)
    
    # Loại bỏ các giá trị trùng lặp
    df = df.drop_duplicates()

    output_file_path = f'RAW_data.xlsx'  
    df.to_excel(output_file_path, index=False)
    file_path = os.path.join(base_directory, f'RAW_data.xlsx')
    
    ########## Edited Report ############
    # Đọc file Excel vào DataFrame
    df = pd.read_excel(file_path)

    # Loại bỏ cột 'Title Search', 'URL', 'Mô tả' (chắc chắn rằng tên cột đúng)
    columns_to_remove = ['Title Search', 'URL', 'Mô tả']
    df = df.drop(columns=columns_to_remove, errors='ignore')

    # Đổi tên cột 'Status' thành 'Trạng thái backlink URLs'
    df = df.rename(columns={'Status': 'Trạng thái backlink URLs'})

    # Thêm cột mới 'Số lượng' để đếm số lượng giá trị trùng lặp trong cột A
    df['Số lượng'] = df.groupby('Domain')['Domain'].transform('count')

    # Loại bỏ các giá trị trùng lặp trong cột A
    df = df.drop_duplicates(subset='Domain', keep='first')

    # Kiểm tra và thêm giá trị "N/A" cho các ô trống trong cột 'Phương thức'
    df['Phương thức'] = df['Phương thức'].fillna('N/A')

    # Lưu DataFrame đã được chỉnh sửa vào một file Excel mới
    output_file_path = f'[TTCNTT - HISSC] Backlink_Edited_report_{current_datetime}.xlsx'  
    df.to_excel(output_file_path, index=False)

    print("Các giá trị trùng lặp trong cột A đã được xóa và số lượng bị trùng đã được tính. Lưu vào file Excel mới.")


    ############## RAW REPORT #############
    workbook = openpyxl.load_workbook(file_path)
    # Chọn sheet cần làm việc
    sheet = workbook.active
    # Định dạng dòng 1 từ cột A đến F
    for column in range(1, 7):  # Cột A đến F là cột từ 1 đến 6
        cell = sheet.cell(row=1, column=column)

        # Đặt kiểu cột là "Neutral"
        cell.style = "Neutral"

        # Đặt màu chữ mà bạn muốn (màu đen, mã màu là "9C6500")
        cell.font = Font(bold=True, color="9C6500", size=13)  # In đậm chữ và màu 9C6500

        # Align dọc giữa (middle align)
        cell.alignment = Alignment(vertical="center", horizontal="center")

        # Đặt all border cho ô
        border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        cell.border = border

    # Đặt cứng giá trị chiều cao là 20
    fixed_row_height = 24

    # Đặt chiều cao của dòng 1
    sheet.row_dimensions[1].height = fixed_row_height

    # Tự động giãn ra chiều rộng của cột dựa trên độ dài giá trị trong cột
    for column in range(1, 7):  # Cột A đến F là cột từ 1 đến 7
        if 2 <= column <= 4:  # Không áp dụng giãn chiều rộng cho cột 2 đến 5
            continue

        max_length = max(len(str(cell.value)) for cell in sheet[get_column_letter(column)])
        adjusted_width = (max_length + 2) * 1  # 1 là hệ số điều chỉnh (tùy chỉnh)
        sheet.column_dimensions[get_column_letter(column)].width = adjusted_width

    # Đặt chiều rộng của cột 2, 3 và 4 là 50
    for column in range(2, 5):  # Cột B đến D là cột từ 2 đến 4
        sheet.column_dimensions[get_column_letter(column)].width = 50

    # Tìm và định dạng các ô trong cột 5
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=5, max_col=5):
        for cell in row:
            if cell.value == "URL còn tồn tại":
                cell.style = "Bad"  # Đặt kiểu 'bad'
                cell.font = Font(bold=True, color="9C0006")  # In đậm với màu 9C0006
            elif cell.value in ["Error 404 Not found", "Đã bị chặn", "Không tồn tại trên Server"]:
                cell.style = "Good"  # Đặt kiểu 'good'
                cell.font = Font(bold=True, color="006100")  # In đậm với màu 006100
            else :
                cell.font = Font(bold=True, color="002060", italic=True)  # In đậm với màu 006100   
                        
    # Tìm và định dạng các ô trong cột 3 ngoài dòng 1
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=3, max_col=3):
        for cell in row:
            cell.font = Font(italic=True, color="0000FF")  # Chuyển sang kiểu italic và màu xanh như màu của đường link

    # Lưu file Excel sau khi đã thay đổi
    workbook.save(file_path)

    # Đóng file Excel
    workbook.close()
    # Đổi tên file cũ thành 'merged_data.xlsx'
    os.rename(file_path, f'[TTCNTT - HISSC] Backlink_Raw_report_{current_datetime}.xlsx')
    
    # Replace 'YOUR_API_TOKEN' with your actual Telegram Bot API token
    api_token = 'api_token'
    # Use your channel chat ID
    chat_id = 'chat_id'  
    # Use your channel thread ID
    message_thread_id = 'thread_id'
    message = f'<b>[TTCNTT - HISSC] Báo cáo tuần Backlink: </b>\n'

    # Lưu DataFrame chung vào tệp Excel chung với mã hóa UTF-8
    file_excel_report_raw = os.path.join(base_directory, f'[TTCNTT - HISSC] Backlink_Raw_report_{current_datetime}.xlsx')
    file_excel_report_edited = os.path.join(base_directory, f'[TTCNTT - HISSC] Backlink_Edited_report_{current_datetime}.xlsx')

    # Create the message data
    data = {
        'chat_id': chat_id,
        'text': message,
        'message_thread_id': message_thread_id,
        'parse_mode': 'HTML'
    }

    data_doc = {
        'caption': 'cap',
        'message_thread_id': message_thread_id,
    }

    files_raw = {
        'document': open(f'{file_excel_report_raw}', 'rb'),
    }
    
    files_edited = {
        'document': open(f'{file_excel_report_edited}', 'rb'),
    }    

    # Send the message using the Telegram Bot API
    url_chat = f'https://api.telegram.org/bot{api_token}/sendMessage'
    response_chat = requests.post(url_chat, data=data)

    url_doc = f'https://api.telegram.org/bot{api_token}/sendDocument?chat_id={chat_id}'
    response_doc_raw = requests.post(url_doc, data=data, files=files_raw)
    response_doc_edited = requests.post(url_doc, data=data, files=files_edited)
    
    if response_doc_raw.status_code == 200: 
        # Đóng file trước khi xóa
        files_raw['document'].close()
        if os.path.exists(file_excel_report_raw):
            os.remove(file_excel_report_raw)
        
    if response_doc_edited.status_code == 200:   
        files_edited['document'].close()
        if os.path.exists(file_excel_report_edited):
            os.remove(file_excel_report_edited)
        
        
file_excel = os.path.join(base_directory, f'results_backlink_{current_datetime}.xlsx')        
if os.path.exists(file_excel):
    os.remove(file_excel)