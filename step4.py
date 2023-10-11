import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import datetime
import os

# Tạo đường dẫn đến thư mục /his-backlink/
base_directory = os.path.abspath("/his-backlink/")

# Mở file Excel
current_date = datetime.datetime.now().strftime('%Y-%m-%d')
file_path = os.path.join(base_directory, f'results_backlink_{current_date}.xlsx')
print(file_path)
workbook = openpyxl.load_workbook(file_path)

# Chọn sheet cần làm việc
sheet = workbook.active

# Định dạng dòng 1 từ cột A đến F
for column in range(1, 7):  # Cột A đến F là cột từ 1 đến 6
    cell = sheet.cell(row=1, column=column)

    # Đặt kiểu cột là "Neutral"
    cell.style = "Neutral"

    # Đặt màu chữ mà bạn muốn (màu đen, mã màu là "9C6500")
    cell.font = Font(bold=True, color="9C6500", size=13)  # In đậm chữ và màu 9C6500

    # Align dọc giữa (middle align)
    cell.alignment = Alignment(vertical="center", horizontal="center")

    # Đặt all border cho ô
    border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    cell.border = border

# Đặt cứng giá trị chiều cao là 20
fixed_row_height = 24

# Đặt chiều cao của dòng 1
sheet.row_dimensions[1].height = fixed_row_height

# Tự động giãn ra chiều rộng của cột dựa trên độ dài giá trị trong cột
for column in range(1, 7):  # Cột A đến F là cột từ 1 đến 7
    if 2 <= column <= 4:  # Không áp dụng giãn chiều rộng cho cột 2 đến 5
        continue

    max_length = max(len(str(cell.value)) for cell in sheet[get_column_letter(column)])
    adjusted_width = (max_length + 2) * 1  # 1 là hệ số điều chỉnh (tùy chỉnh)
    sheet.column_dimensions[get_column_letter(column)].width = adjusted_width

# Đặt chiều rộng của cột 2, 3 và 4 là 50
for column in range(2, 5):  # Cột B đến D là cột từ 2 đến 4
    sheet.column_dimensions[get_column_letter(column)].width = 50

# Tìm và định dạng các ô trong cột 5
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=5, max_col=5):
    for cell in row:
        if cell.value == "Còn tồn tại":
            cell.style = "Bad"  # Đặt kiểu 'bad'
            cell.font = Font(bold=True, color="9C0006")  # In đậm với màu 9C0006
        elif cell.value in ["Error 404 Not found", "Đã bị chặn", "Không tồn tại trên Server"]:
            cell.style = "Good"  # Đặt kiểu 'good'
            cell.font = Font(bold=True, color="006100")  # In đậm với màu 006100
        else :
            cell.font = Font(bold=True, color="002060", italic=True)  # In đậm với màu 006100   
                     
# Tìm và định dạng các ô trong cột 3 ngoài dòng 1
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=3, max_col=3):
    for cell in row:
        cell.font = Font(italic=True, color="0000FF")  # Chuyển sang kiểu italic và màu xanh như màu của đường link


# Lưu file Excel sau khi đã thay đổi
workbook.save(file_path)

# Đóng file Excel
workbook.close()
